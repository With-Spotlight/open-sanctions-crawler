import re
from openpyxl import load_workbook
from rigour.mime.types import XLSX
from typing import Optional

from zavod import Context, helpers as h

SOURCE_URLS = ["https://aml.iq/?page_id=2169", "https://aml.iq/?page_id=2171"]

LOC_PERSON_LISTS = [
    "2017",
    "2018",
    "2019",
    "2020",
    "2021",
    "2022",
    "2023",
    "2024",
    "2025",
]
LOC_ENT_LIST = ["الكيانات"]
LOC_IGNORE_LIST = ["الافراد", "القوائم المحلية "]
INT_LISTS = ["كيانات", "افراد"]

YEAR_PATTERN = re.compile(r"\b(" + "|".join(LOC_PERSON_LISTS) + r")\b")
ENTITY_NAME_REASON = re.compile(r"\s*للتوسط ببيع وشراء العملات الاجنبية$")


def clean_entity_name(entity_name: str) -> Optional[str]:
    if entity_name:
        match = ENTITY_NAME_REASON.search(entity_name)
        if match:
            return ENTITY_NAME_REASON.sub("", entity_name).strip()
    return entity_name


def extract_reason(entity_name: str) -> Optional[str]:
    """Extracts the reason for listing from the entity name."""
    if entity_name is None:
        return None
    match = ENTITY_NAME_REASON.search(entity_name)
    return match.group().strip() if match else None


def extract_listing_date(decision_number: str) -> Optional[str]:
    """Extracts the listing year from the decision number."""
    if decision_number is None:
        return None
    match = YEAR_PATTERN.search(decision_number)
    return match.group(0) if match else None


def crawl_row(row: dict, context: Context):
    entity_name = row.pop("asm_alkyan", None)  # اسم الكيان
    id = row.pop("t")  # ت
    decision_number = row.pop("rqm_alqrar")  # رقم القرار

    reason = extract_reason(entity_name)
    entity_name = clean_entity_name(entity_name)
    listing_date = extract_listing_date(decision_number)

    if entity_name:
        entity = context.make("LegalEntity")
        entity.id = context.make_id(id, entity_name, decision_number)
        entity.add("topics", "debarment")
        entity.add("name", entity_name, lang="ara")
        context.emit(entity)
        sanction = h.make_sanction(context, entity)
    else:
        name = row.pop("asm_alshkhs", row.pop("asma_alashkhas", None))  # اسم الشخص
        person = context.make("Person")
        person.id = context.make_id(id, name)
        person.add("topics", "debarment")
        person.add("nationality", row.pop("aljnsyt", None), lang="ara")  # الجنسية
        h.apply_date(person, "birthDate", row.pop("altwld"))  # التولد
        h.apply_name(
            person,
            full=name,
            matronymic=row.pop("asm_alam"),  # اسم الام
            lang="ara",
        )
        context.emit(person)
        sanction = h.make_sanction(context, person)

    sanction.add("recordId", decision_number)
    sanction.add("reason", reason, lang="ara")
    h.apply_date(sanction, "listingDate", listing_date)
    context.emit(sanction)

    context.audit_data(row)


def process_xlsx(
    context,
    url: str,
    filename: str,
    expected_sheets: list,
    title: str,
    ignore_sheets: list = [],
):
    doc = context.fetch_html(url, cache_days=1)
    link = doc.xpath(f'//article[@id="post-{url.split("=")[-1]}"]//a/@href')

    assert len(link) == 1, link
    file_url = link[0]
    assert file_url.endswith(".xlsx"), file_url
    assert title in file_url, file_url

    path = context.fetch_resource(filename, file_url)
    context.export_resource(path, XLSX)

    wb = load_workbook(path, read_only=True)
    for sheet in expected_sheets:
        for row in h.parse_xlsx_sheet(context, wb[sheet], skiprows=3):
            crawl_row(row, context)

    assert set(wb.sheetnames) == set(expected_sheets + ignore_sheets)


def crawl(context: Context):
    process_xlsx(
        context,
        SOURCE_URLS[0],
        "international.xlsx",
        INT_LISTS,
        "القائمة-الدولية",  # International list
    )
    process_xlsx(
        context,
        SOURCE_URLS[1],
        "local.xlsx",
        LOC_PERSON_LISTS + LOC_ENT_LIST,
        "القوائم-المحلية",  # Local lists
        LOC_IGNORE_LIST,
    )
